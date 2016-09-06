import matplotlib.pyplot as plt
import numpy
import re
from openpyxl import load_workbook
import sys

#plots the frequency of expression values in a xlsx file
#python expression Repartition.py path_to_file sheet_name first_cell:last_cell

all_expression={}
wb=load_workbook(sys.argv[1])
sheet=wb['Culture']
for line in sheet.iter_rows("C2:C8331"):
	if unicode(line[0].value) in all_expression.keys():
		all_expression[unicode(line[0].value)]+=1
	else:
		all_expression[unicode(line[0].value)]=1

values=[]
for i in all_expression.keys():
	if not i in values:
		values.append(int(i))
tab_values=sorted(values)

freq=[]
for i in tab_values:
	freq.append(all_expression[str(i)])

plt.subplots()
plt.plot(tab_values,freq)
plt.show()
