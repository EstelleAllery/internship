
import re
from openpyxl import load_workbook
import math
import sys

#Calculates RPKM for each cell in a xslx file containing read counts
#python rpkm.py path_to_xslx_file path_to_output path_to_gene_length_file sheets(separated by spaces)

normExp_file=open(sys.argv[2],"w")

expression_file=load_workbook(sys.argv[1])
culture_data=expression_file["Culture"]
planta_data=expression_file["Planta"]

expression=""
totalReads1=0
totalReads2=0
totalReads3=0
totalReads4=0
totalReads5=0
totalReads6=0
totalReads7=0
totalReads8=0

for i in culture_data.iter_rows('A2:E8331'):
	totalReads1+=float(i[1].value)
	totalReads2+=float(i[2].value)
	totalReads3+=float(i[3].value)
	totalReads4+=float(i[4].value)

for i in planta_data.iter_rows('A2:E8331'):
	totalReads5+=float(i[1].value)
	totalReads6+=float(i[2].value)
	totalReads7+=float(i[3].value)
	totalReads8+=float(i[4].value)

totalReads=[totalReads1,totalReads2,totalReads3,totalReads4,totalReads5,totalReads6,totalReads7,totalReads8]
print totalReads
normExp1=0.0
normExp2=0.0

# RPKM =   numberOfReads / ( geneLength/1000 * totalNumberOfReads/1,000,000 )
count=0
row=1
for i in culture_data.iter_rows('A2:E8331'):
	row+=1
	normExp=[]
	with open(sys.argv[3]) as geneLengthFile:
		for line in geneLengthFile:
			if re.search(i[0].value,line):
				count+=1
				content=line.split("\t")
				if int(content[1])!=0:
					for j in range(1,5):
						normExp.append(float(i[j].value)*1000000000/totalReads[j-1]/float(content[1]))
					for k in planta_data.iter_rows('A'+str(row)+':E'+str(row)):
						if k[0].value==i[0].value:
							for j in range(5,9):
								normExp.append(float(k[j-4].value)*1000000000/totalReads[j-1]/float(content[1]))
					geneName=content[0]
					normExp_file.write(geneName)
					for n in range(0,len(normExp)):
						normExp_file.write("\t")
						normExp_file.write(str(normExp[i]))
					normExp_file.write("\n")
				else:
					print line
					print i[0].value
				break

print count
normExp_file.close()
